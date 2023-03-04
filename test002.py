#!/usr/bin/python
# -*- coding: UTF-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook

"""
加载已存在的工作簿
"""
#实例化
wb=Workbook()
wb = load_workbook('/Users/gouchaonan/Desktop/abcd.xlsx') # openpyxl第三方库只能处理.xlsx格式的Excel表格
print(wb)
#获取worksheet对象
ws=wb.active
print(ws)

"""
新建名为XXX的工作簿
"""
b = Workbook('abc.xlsx')    #新建名为abc的工作簿
wb.save('abc.xlsx')         #保存工作簿，完成新工作簿的建立(将覆盖同名文件且无警告）

#打开excel文件
print(wb.sheetnames)  #打印工作的表的名字
print(ws['A1'].value)  # 获取指定位置的单元格对象

"""
修改当前所处工作表名称，【注意】对工作簿修改后一定要记得执行保存操作！
"""
ws.title='MS'
wb.save('/Users/gouchaonan/Desktop/abcd.xlsx')

print(ws.values, end='\n\n')  # 生成器对象，将一行单元格作为元组单元--》组成的生成器
print(list(ws.values), end='\n\n') # 将生成器对象转换为列表数据，列表中是生成器中的所有数据

for i in ws.values:
    print(i)


for i in ws.iter_rows(min_col=1, max_col=3, min_row=1, max_row=10):
    print(i)

print(ws.iter_rows())  # 将每一个单元格对象（按行数据中的单元格）作为一个元组单元--》组成的生成器
print(ws.rows, end='\n\n')  # 将每一个单元格对象（按行数据中的单元格）作为一个元组单元--》组成的生成器

print(list(ws.rows), end='\n\n')  # 列表中存放的是生成器中的每一个单元格对象

for i in ws.rows:  # 遍历以单元格对象为单位的元组单元组成的生成器
    for j in i:  # 遍历单元格对象中的属性
        print(j, j.value)